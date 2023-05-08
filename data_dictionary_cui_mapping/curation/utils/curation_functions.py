"""

This file contains custom functions used in scripts designed to merge curated UMLS CUIs back with the original examples dictionary file.

"""

from functools import reduce
from pathlib import Path
import numpy as np
import pandas as pd
from openpyxl.utils import get_column_letter
from prefect import flow, task
from prefect.task_runners import SequentialTaskRunner

from data_dictionary_cui_mapping.utils import helper as helper


@task(name="Adding search_ID column")
def add_search_ID_col(df):
    """Add search ID column"""

    col_search_ID = "search_ID"
    df.insert(
        1, col_search_ID, range(1, 1 + len(df))
    )  # insert column search_ID as number 1 through length of df_curation
    return df


@task(name="Adding search pipeline name column")
def add_search_pipeline_col(df, pipeline_name):
    df.insert(
        1, "pipeline_name", pipeline_name
    )  # insert column pipeline_name e.g., metamp, umls, semantic_search
    return df


@task(name="Subsetting dataframe with curation related columns")
def curation_cols_filter(df, df_dd, cols_curation: list):
    """Create subset of dataframe with columns necessary for curation"""

    rmv_cols = list(set(df_dd.columns).difference(cols_curation))
    df = df.drop(rmv_cols, axis=1).copy()  # create curation file
    return df


@flow(flow_run_name="Formatting dataframe for curation")
def format_curation_dataframe(df_dd, df_dd_preprocessed, pipeline_name, cfg):
    """Create dataframe for curation"""

    df_curation = (
        df_dd_preprocessed.copy()
        .pipe(
            curation_cols_filter,
            df_dd,
            cfg.custom.curation_settings.information_columns,
        )
        .pipe(  # use cheatsheet # TODO: need to implement this in the future
            add_search_ID_col
        )
        .pipe(add_search_pipeline_col, pipeline_name)
    )
    # query_terms_cols = [col for col in df_curation.columns if re.search(r'query_term_\d+', col)]  # find columns in df_curation that match query_term_ then any number
    return df_curation  # , query_terms_cols


@task(name="Creating curation file")
def create_curation_file(
    dir_step1, df_dd, df_dd_preprocessed, df_curation, df_results, cfg
):
    """Create curation file"""

    df_results = df_results.replace(r"^\s*$", np.nan, regex=True)
    df_final = pd.merge(
        df_curation,
        df_results,
        left_on="search_ID",
        right_on="search_ID",
        how="outer",
        suffixes=("", "_y"),
    )
    df_final.drop(df_final.filter(regex="_y$").columns, axis=1, inplace=True)
    df_final["keep"] = np.nan
    if cfg.custom.curation_settings.file_settings.excel.order_cols_curation:
        df_final = reorder_cols(
            df_final,
            cfg.custom.curation_settings.file_settings.excel.order_cols_curation,
        )
    fp_prefix = cfg.custom.curation_settings.file_settings.file_prefix
    fp_step1 = dir_step1 / f"{fp_prefix}_Step-1_curation_keepCol.xlsx"
    with pd.ExcelWriter(fp_step1, engine="openpyxl") as writer:
        df_final.to_excel(
            writer,
            sheet_name=cfg.custom.curation_settings.file_settings.excel.sheet_names.sheet1,
            index=False,
        )
        df_dd.to_excel(
            writer,
            sheet_name=cfg.custom.curation_settings.file_settings.excel.sheet_names.sheet2,
            index=False,
        )
        df_dd_preprocessed.to_excel(
            writer,
            sheet_name=cfg.custom.curation_settings.file_settings.excel.sheet_names.sheet3,
            index=False,
        )

        ws1 = writer.sheets[
            cfg.custom.curation_settings.file_settings.excel.sheet_names.sheet1
        ]
        set_col_widths(ws1, df_final)
        set_hidden_cols(
            ws1,
            df_final,
            cfg.custom.curation_settings.file_settings.excel.hide_cols_curation,
        )

    return df_final


def get_curation_excel_file(prompt: str):
    """Get curation Excel file path"""

    fp_curation = helper.choose_file(prompt)
    return fp_curation


@task(name="Loading curation excel file")
def load_curation_excel_file(fp_curation: str, cfg):
    """Load curation Excel file"""

    df_UMLS_curation = pd.read_excel(
        fp_curation,
        sheet_name=cfg.custom.curation_settings.file_settings.excel.sheet_names.sheet1,
        header=0,
        keep_default_na=False,
    )
    df_Data_Dictionary = pd.read_excel(
        fp_curation,
        sheet_name=cfg.custom.curation_settings.file_settings.excel.sheet_names.sheet2,
        header=0,
        keep_default_na=False,
    )
    df_Data_Dictionary_extracted = pd.read_excel(
        fp_curation,
        sheet_name=cfg.custom.curation_settings.file_settings.excel.sheet_names.sheet3,
        header=0,
        keep_default_na=False,
    )
    return df_UMLS_curation, df_Data_Dictionary, df_Data_Dictionary_extracted


@task(name="Filtering rows by keep column")
def filter_keep_col(df):
    """Filter out rows where keep column is empty"""

    df["keep"].replace(
        r"^\s*$", np.nan, regex=True, inplace=True
    )  # makes sure any empty spaces are nan and won't be removed in next step
    df = df[df["keep"].notnull()].reset_index(
        drop=True
    )  # returns df_UMLS_curation with only rows where keep is not null
    return df


@task(name="Reordering rows based on keep column number/letter")
def order_keep_col(df):
    """Orders rows in keep column by number and letter e.g., 1a, 1b, 2a, 2b, 3a, 3b"""

    df["keep"] = df["keep"].astype(str)
    df["keep_num"] = [x[0] for x in df["keep"]]
    df["keep_letter"] = [x[1:] if len(x) > 1 else "" for x in df["keep"]]
    df = df.sort_values(by=["keep", "keep_letter"])
    df["keep_concat"] = ["concat" if len(x) > 1 else "" for x in df["keep"]]
    return df


@task(name="Concatenating multi-CUI concepts")
def concat_mult_cuis(df, cols_join_on, umls_columns):
    """Concatenate multi-CUI concept by /"""

    df = df.groupby(cols_join_on + ["keep_concat", "keep_num"])[
        umls_columns + ["keep_num"]
    ].agg(
        "/".join
    )  # TODO need to deal with logic of missing PVDs
    return df


@task(name="Merging curated cuis with original examples dictionary")
def merge_with_dictionary(
    df_left,
    df_right,
    how: str,
    cols_join_on,
    suffixes: tuple,
):  # TODO maybe add validate check
    """Merge curation file with examples dictionary"""

    df = pd.merge(
        df_left,
        df_right,
        how=how,
        left_on=cols_join_on,
        right_on=cols_join_on,
        suffixes=suffixes,
    )
    return df


@task(name="Concatenating multiple cuis")
def concat_cols_umls(df, umls_columns: list):
    """Concatenate multiple CUIs in one row"""

    df_dict = {}
    for col in umls_columns:
        df_dict[col] = (
            df.groupby(["variable name"])[col].agg("|".join).reset_index(name=col)
        )  # TODO need to deal hard-coded pv col here
    df_merged = reduce(
        lambda left, right: pd.merge(left, right, on=["variable name"]),
        df_dict.values(),
    )
    return df_merged


# @task(name="Reordering examples dictionary columns")
def reorder_cols(df, order: list):
    """Reorder columns"""

    df = df[order]
    return df


@task(name="Manual override of column values")
def override_cols(df, override: dict):
    """Custom function to accommodate current bug in BRICS examples dictionary import process that wants multi-CUI concepts to have a single source terminology
    e.g., C123456/C234567 -> UMLS instead of UMLS/UMLS"""

    sep = override.sep
    cols = override.columns
    value = override.value
    for col in cols:
        temp1 = list(df[col].str.split(sep))
        temp2 = list(map(lambda x: [value for val in x if len(val) > 0], temp1))
        temp3 = list(map(lambda x: sep.join(x), temp2))
        # set col in df to temp3 without setting value on a copy of a slice from DataFrame
        df.loc[:, col] = temp3
    return df


# EXCEL FORMATTING


# @task(name="Getting excel columns max width")
def get_col_max_widths(df):
    """Used to set excel column width to maximum character length in column"""

    # First we find the maximum length of the index column
    idx_max_width = max(
        [len(str(s)) for s in df.index.values] + [len(str(df.index.name))]
    )
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return [
        idx_max_width + max([len(str(s)) for s in df[col].values] + [len(col)])
        for col in df.columns
    ]


# @task(name="Getting excel columns header widths")
def get_header_widths(df):
    """Used to set excel column width to character length of column header"""

    # First we find the maximum length of the index column
    idx_header_width = [len(str(s)) + 4 for s in df.columns]
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return idx_header_width


# @task(name="Setting excel column widths")
def set_col_widths(ws, df):
    """Set column widths in excel"""

    col_widths = get_header_widths(df)
    for i, width in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width


# @task(name="Setting hidden excel columns")
def set_hidden_cols(ws, df, hidden_cols: list):
    """Set hidden columns in excel"""

    if hidden_cols:
        cols_hide_idx = [df.columns.get_loc(c) for c in hidden_cols if c in df]
        for col_idx in cols_hide_idx:
            col_letter_hide = get_column_letter(col_idx + 1)
            # these columns are hidden as they aren't useful for review
            ws.column_dimensions[col_letter_hide].hidden = True
