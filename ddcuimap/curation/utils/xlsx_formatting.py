"""

Functions for formatting Excel curation file.

"""

from openpyxl.utils import get_column_letter

from ddcuimap.curation import log


# EXCEL FORMATTING


@log(msg="Getting excel columns max width")
def get_col_max_widths(df):
    """Used to set excel column width to maximum character length in column"""

    # First we find the maximum length of the index column
    idx_max_width = max(
        [len(str(s)) for s in df.index.values] + [len(str(df.index.name))]
    )
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return [
        idx_max_width + max([len(str(s)) for s in df[col].values] + [len(col)])
        for col in df.columns
    ]


@log(msg="Getting excel columns header widths")
def get_header_widths(df):
    """Used to set excel column width to character length of column header"""

    # First we find the maximum length of the index column
    idx_header_width = [len(str(s)) + 4 for s in df.columns]
    # Then, we concatenate this to the max of the lengths of column name and its values for each column, left to right
    return idx_header_width


@log(msg="Setting excel column widths")
def set_col_widths(ws, df):
    """Set column widths in excel"""

    col_widths = get_header_widths(df)
    for i, width in enumerate(col_widths):
        ws.column_dimensions[get_column_letter(i + 1)].width = width


@log(msg="Setting hidden excel columns")
def set_hidden_cols(ws, df, hidden_cols: list):
    """Set hidden columns in Excel"""

    if hidden_cols:
        cols_hide_idx = [df.columns.get_loc(c) for c in hidden_cols if c in df]
        for col_idx in cols_hide_idx:
            col_letter_hide = get_column_letter(col_idx + 1)
            # these columns are hidden as they aren't useful for review
            ws.column_dimensions[col_letter_hide].hidden = True
